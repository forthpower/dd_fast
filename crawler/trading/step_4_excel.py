import os
import csv
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(__file__)
CRYPTO_EXCEL_PATH = os.path.join(BASE_DIR, "crypto.xlsx")
CSV_DIR = os.path.join(BASE_DIR, "result", "csv")
SHEET_NAME = 'Strategy Watchlist'


def load_nov_numbers_from_csv() -> dict:
    """从最新的 CSV 文件中读取 URL -> Nov 数字的字典"""
    if not os.path.exists(CSV_DIR):
        print(f"❌ CSV 目录不存在: {CSV_DIR}")
        return {}
    
    files = [f for f in os.listdir(CSV_DIR) if 'last_number' in f and f.endswith(".csv")]
    if not files:
        print(f"❌ 未找到包含 'last_number' 的 CSV 文件")
        return {}
    
    latest_file = max(files, key=lambda f: os.path.getmtime(os.path.join(CSV_DIR, f)))
    latest_path = os.path.join(CSV_DIR, latest_file)
    
    url_to_number = {}
    try:
        with open(latest_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader, None)  # 跳过表头
            for row in reader:
                if len(row) >= 2 and row[0].strip() and row[1].strip():
                    url_to_number[row[0].strip()] = row[1].strip()
        print(f"✅ 从 {latest_file} 加载 {len(url_to_number)} 条数据")
    except Exception as e:
        print(f"❌ 读取 CSV 失败: {e}")
        return {}
    
    return url_to_number


def update_excel_with_nov_numbers():
    """读取 CSV 数据，根据 LINK 列匹配，将 Nov 数字填入 NOV 列"""
    # 1. 加载 CSV 数据为字典
    url_to_number = load_nov_numbers_from_csv()
    if not url_to_number:
        print("❌ 没有可用的 Nov 数据")
        return
    
    # 2. 读取 Excel
    if not os.path.exists(CRYPTO_EXCEL_PATH):
        print(f"❌ Excel 文件不存在: {CRYPTO_EXCEL_PATH}")
        return
    
    # 2. 使用 openpyxl 直接操作 Excel，保留所有格式和内容
    try:
        wb = load_workbook(CRYPTO_EXCEL_PATH)
        if SHEET_NAME not in wb.sheetnames:
            print(f"❌ Sheet '{SHEET_NAME}' 不存在")
            return
        
        ws = wb[SHEET_NAME]
        
        # 3. 自动查找表头行（查找包含 "LINK" 的行）
        header_row = None
        link_col_idx = None
        nov_col_idx = None
        
        for row_idx in range(1, min(ws.max_row + 1, 20)):  # 在前20行中查找表头
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                cell_value = str(cell.value or '').upper()
                if 'LINK' in cell_value:
                    header_row = row_idx
                    link_col_idx = col_idx
                    break
            if header_row is not None:
                break
        
        if header_row is None or link_col_idx is None:
            print("❌ 未找到包含 LINK 的表头行")
            return
        
        print(f"✅ 找到表头行: 第 {header_row} 行")
        
        # 4. 在表头行查找 NOV 列
        for col_idx, cell in enumerate(ws[header_row], start=1):
            cell_value = str(cell.value or '').upper()
            if cell_value == 'NOV':
                nov_col_idx = col_idx
                break
        
        # 5. 如果 NOV 列不存在，在表头行创建
        if nov_col_idx is None:
            # 找到最后一列，在下一列添加 NOV
            max_col = ws.max_column
            nov_col_idx = max_col + 1
            ws.cell(row=header_row, column=nov_col_idx, value='NOV')
            print(f"✅ 创建 NOV 列（列 {nov_col_idx}）")
        else:
            print(f"✅ 找到 NOV 列（列 {nov_col_idx}）")
        
        # 6. 更新数据行（从表头行的下一行开始）
        updated_count = 0
        for row_idx in range(header_row + 1, ws.max_row + 1):
            link_cell = ws.cell(row=row_idx, column=link_col_idx)
            link_value = str(link_cell.value or '').strip()
            
            if link_value and link_value in url_to_number:
                nov_value = url_to_number[link_value]
                nov_cell = ws.cell(row=row_idx, column=nov_col_idx)
                nov_cell.value = nov_value
                updated_count += 1
                print(f"✅ [行 {row_idx}] {link_value[:50]}... -> {nov_value}")
        
        # 7. 保存文件（保留所有格式、公式、其他 sheet 等）
        wb.save(CRYPTO_EXCEL_PATH)
        wb.close()
        print(f"\n✅ 更新完成: {updated_count} 条数据已填入 NOV 列")
        print(f"📄 已保存: {CRYPTO_EXCEL_PATH}")
        
    except Exception as e:
        print(f"❌ 处理 Excel 失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    update_excel_with_nov_numbers()